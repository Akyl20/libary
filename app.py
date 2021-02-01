from flask import Flask, render_template
from openpyxl import load_workbook
app = Flask(__name__)
@app.route("/")
def homepage():
    return "Тест!"
@app.route("/books/")
def books():
    excel = load_workbook("tales.xlsx")
    page = excel["Лист1"]
    
    tales = [tale.value for tale in page ["A"]][1:]

    authors = [author.value for author in page ["B"]][1:]
   
    html = "<h1>Туд будет список книг :</h1>"
        

    for i in range(len(tales)):
        html += f"<h2>{tales[i]} - {authors[i]}</h2>"
    return html    

@app.route("/authors/")
def authors():
    excel = load_workbook("tales.xlsx")
    page = excel["Лист1"]
    authors = {author.value for author in page["B"][1:]}
    return render_template("authors.html")